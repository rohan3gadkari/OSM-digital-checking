import os
from flask import Flask, render_template, request, jsonify, redirect, url_for, session, flash, send_file
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

app = Flask(__name__)
app.secret_key = 'your_super_secret_secure_key_here'

# 📂 आन्सर शीट फोल्डर पाथ
UPLOAD_FOLDER = os.path.join('static', 'answer_sheets')
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

# 📊 मास्टर स्ट्रक्चर: Q.1 चे १४ उपप्रश्न (Max: 2) आणि Q.2 ते Q.7 मुख्य प्रश्न (Max: 5)
GLOBAL_STRUCTURE = []
for i in range(1, 15):
    GLOBAL_STRUCTURE.append({"id": f"q1_{i}", "label": f"Q.1 ({i})", "max": 2})
for i in range(2, 8):
    GLOBAL_STRUCTURE.append({"id": f"q{i}", "label": f"Q.{i}", "max": 5})

# 🗄️ डेटाबेस आणि परमनंट लॉक सिस्टीम (मेमरी रिपॉझिटरी)
EVALUATION_DATABASE = []
LOCKED_PAPERS = set()  # एकदा सेव्ह झालेले बारकोड्स इथे कायमचे लॉक होतील

# युझर क्रेडेंशियल्स
USER_DATA = {"Prof_Gadkari": "password123"}

@app.route('/')
def index():
    if 'user' not in session:
        return redirect(url_for('login'))
        
    current_paper = request.args.get('paper', 'sample.pdf')
    total_max_marks = sum(q['max'] for q in GLOBAL_STRUCTURE)

    # प्रत्येक पेपरसाठी स्वतंत्र सॅम्पल बारकोड मॅपिंग
    paper_barcode_map = {
        "sample.pdf": "BAR1009238",
        "student_02.pdf": "BAR9920182",
        "student_03.pdf": "BAR5549210"
    }
    current_barcode = paper_barcode_map.get(current_paper, "BAR_DYNAMIC_01")

    # 🔒 पेपर आधीच लॉक झाला आहे का ते तपासा ("true" / "false")
    is_locked = "true" if current_barcode in LOCKED_PAPERS else "false"

    return render_template('dashboard.html', 
                           user=session['user'], 
                           current_paper=current_paper, 
                           structure=GLOBAL_STRUCTURE,
                           total_max=total_max_marks,
                           barcode=current_barcode,
                           is_locked=is_locked)

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        if username in USER_DATA and USER_DATA[username] == password:
            session['user'] = username
            return redirect(url_for('index'))
        else:
            flash("चुकीचा Username किंवा Password!")
            return redirect(url_for('login'))
    return render_template('login.html')

@app.route('/submit_marks', methods=['POST'])
def submit_marks():
    if 'user' not in session:
        return jsonify({'status': 'error', 'message': 'Unauthorized'}), 401
    
    data = request.get_json()
    barcode = data['barcode']

    # 🛑 सिक्युरिटी चेक: पेपर आधीच लॉक असेल तर सबमिशन नाकारा
    if barcode in LOCKED_PAPERS:
        return jsonify({'status': 'error', 'message': 'हा पेपर आधीच परमनंट लॉक झाला आहे!'}), 400

    # मॅन्युअली बदललेले मॅक्स मार्क्स अपडेट करणे
    incoming_max = data.get('max_limits', {})
    for q in GLOBAL_STRUCTURE:
        if q['id'] in incoming_max:
            q['max'] = int(incoming_max[q['id']])

    # डेटाबेसमध्ये परमनंट नोंद (Permanent Save)
    EVALUATION_DATABASE.append({
        "barcode": barcode,
        "paper": data['paper'],
        "examiner": session['user'],
        "scores": data['scores'],
        "max_limits": incoming_max
    })
    
    # 🔐 बारकोड कायमचा लॉक यादीत टाका
    LOCKED_PAPERS.add(barcode)
    print(f"🔒 Paper Locked Successfully in Backend: {barcode}")
    return jsonify({'status': 'success'})

@app.route('/download_excel')
def download_excel():
    if 'user' not in session:
        return "Unauthorized", 401

    # एक्सेल फाईल तयार करणे
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Final Locked Records"
    ws.views.sheetView[0].showGridLines = True  # ग्रिडलाईन्स दाखवा

    # स्टाईल्स व्याख्या
    font_title = Font(name="Segoe UI", size=15, bold=True, color="1E293B")
    font_header = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
    font_data = Font(name="Segoe UI", size=11)
    fill_header = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    border_thin = Border(left=Side(style='thin', color='E2E8F0'), right=Side(style='thin', color='E2E8F0'),
                         top=Side(style='thin', color='E2E8F0'), bottom=Side(style='thin', color='E2E8F0'))

    # टायटल रो
    ws['A1'] = "OSM Master Production Database — Permanently Saved Marks"
    ws['A1'].font = font_title
    ws.row_dimensions[1].height = 25

    # एक्सेल हेडर कॉलम्स रचना
    headers = ["BarCode No.", "Paper File", "Examiner ID", "Lock Status"]
    for q in GLOBAL_STRUCTURE:
        headers.append(f"{q['label']}\n[Max: {q['max']}]")
    headers.append("Grand Total")

    ws.row_dimensions[3].height = 32
    for col_idx, text in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_idx, value=text)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border_thin

    # डेटा रोज भरण्यास सुरुवात
    current_row = 4
    for record in EVALUATION_DATABASE:
        ws.row_dimensions[current_row].height = 20
        sc = record['scores']
        
        # मुख्य माहिती भरणे
        ws.cell(row=current_row, column=1, value=record['barcode']).alignment = Alignment(horizontal="center")
        ws.cell(row=current_row, column=2, value=record['paper']).alignment = Alignment(horizontal="left")
        ws.cell(row=current_row, column=3, value=record['examiner']).alignment = Alignment(horizontal="center")
        
        status_cell = ws.cell(row=current_row, column=4, value="LOCKED & VERIFIED")
        status_cell.font = Font(name="Segoe UI", size=10, bold=True, color="991B1B")
        status_cell.alignment = Alignment(horizontal="center")
        
        # सर्व २० उपप्रश्नांचे प्राप्त गुण भरणे (Column E पासून पुढे)
        c_idx = 5
        for q in GLOBAL_STRUCTURE:
            ws.cell(row=current_row, column=c_idx, value=sc.get(q['id'], 0))
            c_idx += 1
            
        # डायनॅमिक एक्सल SUM फॉर्म्युला लावणे (E{row} ते अंतिम प्रश्नाचा कॉलम)
        last_col_letter = get_column_letter(c_idx - 1)
        total_formula = f"=SUM(E{current_row}:{last_col_letter}{current_row})"
        
        total_cell = ws.cell(row=current_row, column=c_idx, value=total_formula)
        total_cell.font = Font(name="Segoe UI", bold=True, color="065F46")
        total_cell.fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
        
        # स्टाईल्स आणि बॉर्डर्स लागू करणे
        for c in range(1, c_idx + 1):
            cell = ws.cell(row=current_row, column=c)
            if c >= 5 or c in [1, 4]: cell.alignment = Alignment(horizontal="center")
            if c != c_idx and c != 4: cell.font = font_data
            cell.border = border_thin
            
        current_row += 1

    # कॉलम विड्थ (रुंदी) ऑटो-ऍडजस्ट करणे
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = 11
    ws.column_dimensions['A'].width = 16
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['D'].width = 18

    excel_path = "OSM_Evaluation_Sheet.xlsx"
    wb.save(excel_path)
    return send_file(excel_path, as_attachment=True, download_name="OSM_Evaluation_Sheet.xlsx")

if __name__ == '__main__':
    app.run(debug=True)
